
import openpyxl
from openpyxl.styles import Font
import sys

# Take parameters from the command line, if any
# The dimension of the multiplication table is defaulted 19
table_dimension = int(sys.argv[1]) if len(sys.argv)>=2 else 19

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Multiplication Table'

# set label font
font = Font(bold=True, size=12)

# Add the labels on the top and on the left
for i in range(2, 2+table_dimension):
    cell_top = ws.cell(row=1, column=i)
    cell_left = ws.cell(row=i, column=1)
    cell_top.font = cell_left.font = font
    cell_top.value = cell_left.value = i-1

# Add the contents of the multiplication table
for i in range(2, 2+table_dimension):
    for j in range(2, 2+table_dimension):
        ws.cell(row=i, column=j).value = (i-1)*(j-1)

wb.save('MultiplicationTable.xlsx')