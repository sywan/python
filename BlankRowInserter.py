import openpyxl
import sys

"""
if len(sys.argv)<4:
    sys.exit('Format: >>> py BlankRowInsrter.py row_to_insert number_of_rows_to_be_inserted file_to_be_inserted')
else:
    row_to_insert = int(sys.arg[1])
    number_of_rows = int(sys.argv[2])
    file_to_be_inserted = sys.argv[3]
"""

row_to_insert = 3
number_of_rows_to_be_inserted = 2
file_to_be_inserted = 'style.xlsx'

wb = openpyxl.load_workbook(file_to_be_inserted)
ws = wb.active

# Determine number of rows and columns
n_row = ws.max_row
n_col = ws.max_column

sheet_data = []

# read the entire sheet as a list of lists
for i in range(1, n_row+1):
    row_data = []
    for j in range(1, n_col+1):
        row_data.append(ws.cell(row=i, column=j).value)
    sheet_data.append(row_data)

# insert {number_of_rows_to_be_inserted} right above the {row_to_insert}th row
for i in range(number_of_rows_to_be_inserted):
    sheet_data.insert(row_to_insert-1, [None]*n_col)

# Create a sheet and store the result
ws_2 = wb.create_sheet('Blank Line Inserted', 1)
for i in range(1, n_row+1):
    for j in range(1, n_col+1):
        ws_2.cell(row=i, column=j).value = sheet_data[i-1][j-1]

wb.save(file_to_be_inserted)

