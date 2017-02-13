import openpyxl
import sys

"""
if sys.argv[1:]:
    wb = openpyxl.load_workbook(sys.argv[1])
else:
    wb = openpyxl.load_workbook('test.xlsx')
"""
wb = openpyxl.load_workbook('test.xlsx')

ws = wb.active

nRow = ws.max_row
nCol = ws.max_column

sheetData = []

for i in range(1, nRow+1):
    rowData = []
    for j in range(1, nCol+1):
        rowData.append(ws.cell(row=i, column=j).value)
    sheetData.append(rowData)

ws2 = wb.create_sheet('transposed sheet',1)

# Perform Transposition
for i in range(1, nRow+1):
    for j in range(1, nCol+1):
        ws2.cell(row=j, column=i).value = ws.cell(row=i, column=j).value

wb.save('test_out.xlsx')
